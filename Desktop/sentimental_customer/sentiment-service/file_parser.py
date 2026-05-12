import fitz                        # pymupdf — for PDF
from docx import Document          # python-docx — for Word
from openpyxl import load_workbook # openpyxl — for Excel
import os

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".txt", ".csv", ".xlsx"}

def parse_file(filepath: str) -> str:
    """
    Reads a file and returns its full text content as a single string.
    Supports PDF, DOCX, TXT, CSV, XLSX.
    """
    _, ext = os.path.splitext(filepath.lower())

    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported file type: {ext}. Supported: {', '.join(SUPPORTED_EXTENSIONS)}")

    if ext == ".pdf":
        return _parse_pdf(filepath)
    elif ext == ".docx":
        return _parse_docx(filepath)
    elif ext == ".xlsx":
        return _parse_xlsx(filepath)
    elif ext in (".txt", ".csv"):
        return _parse_text(filepath)


def _parse_pdf(filepath: str) -> str:
    doc   = fitz.open(filepath)
    pages = [page.get_text() for page in doc]
    doc.close()
    return "\n".join(pages).strip()


def _parse_docx(filepath: str) -> str:
    doc   = Document(filepath)
    lines = [para.text for para in doc.paragraphs if para.text.strip()]
    return "\n".join(lines).strip()


def _parse_xlsx(filepath: str) -> str:
    wb    = load_workbook(filepath, read_only=True, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            line = " | ".join(str(cell) for cell in row if cell is not None)
            if line.strip():
                lines.append(line)
    wb.close()
    return "\n".join(lines).strip()


def _parse_text(filepath: str) -> str:
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        return f.read().strip()


def split_into_chunks(text: str, max_chars: int = 500) -> list[str]:
    """
    Splits text into sentence-aware chunks for per-chunk analysis.
    Tries to split on sentence boundaries first, then falls back to char limit.
    """
    import re
    # Split on sentence endings
    sentences = re.split(r'(?<=[.!?])\s+', text)
    chunks  = []
    current = ""

    for sentence in sentences:
        if not sentence.strip():
            continue
        if len(current) + len(sentence) <= max_chars:
            current += (" " if current else "") + sentence
        else:
            if current:
                chunks.append(current.strip())
            # If a single sentence exceeds max_chars, hard-split it
            if len(sentence) > max_chars:
                for i in range(0, len(sentence), max_chars):
                    chunks.append(sentence[i:i + max_chars].strip())
            else:
                current = sentence

    if current.strip():
        chunks.append(current.strip())

    return [c for c in chunks if c]